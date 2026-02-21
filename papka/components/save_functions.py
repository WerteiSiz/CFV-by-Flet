import os
import flet as ft

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment

from papka.components.input_fields import Inputs
from papka.components.notifications import show_notification
from papka.components.constants import COLORS

from .calculations import (
    ygol_mesta,
    azimut,
    dalnost,
    poteri,
    factorG,
    atmosphere,
)


# =====================================================
# ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ
# =====================================================

def get_values(inputs: Inputs):
    """Читает значения из Inputs"""

    return dict(
        gradys_d=float(inputs.gradys_d.value),
        gradys_sh=float(inputs.gradys_sh.value),
        tochka=float(inputs.tochka.value),
        chastota=float(inputs.chastota.value),
        diametr=float(inputs.diametr.value),
        k=float(inputs.k.value),
        P=float(inputs.P.value),
        T=float(inputs.T.value),
        p=float(inputs.p.value),
    )


# =====================================================
# TXT
# =====================================================

def save_result_txt(page: ft.Page, inputs: Inputs):

    try:
        v = get_values(inputs)

        angle = ygol_mesta(v["gradys_d"], v["gradys_sh"], v["tochka"])
        az = azimut(v["gradys_d"], v["gradys_sh"], v["tochka"])
        dist = dalnost(v["gradys_d"], v["gradys_sh"], v["tochka"])

        loss = round(poteri(v["chastota"], v["diametr"], v["k"], v["tochka"]), 3)
        gain = round(factorG(v["chastota"], v["k"], v["diametr"]), 3)
        atmos = round(
            atmosphere(
                v["chastota"],
                v["gradys_d"],
                v["gradys_sh"],
                v["tochka"],
                v["T"],
                v["p"],
                v["P"],
            ),
            3,
        )

        result_text = f"""
╔════════════════════════════════╗
║        ОТЧЕТ О РАСЧЕТАХ        ║
╠════════════════════════════════╣
║ Угол места: {angle:.2f}°
║ Азимут: {az:.2f}°
║ Дальность: {dist:.2f} км
║ Потери сигнала: {loss:.2f} дБ
║ Потери атмосферы: {atmos:.2f} дБ
║ Усиление антенны: {gain:.2f} дБ
╚════════════════════════════════╝
"""

        with open("Результаты_Расчётов.txt", "w", encoding="utf-8") as f:
            f.write(result_text)

        show_notification(page, "TXT файл успешно сохранён ✅")

    except ValueError:
        show_notification(page, "Ошибка: проверьте введённые данные", True)


# =====================================================
# WORD
# =====================================================

def save_result_word(page: ft.Page, inputs: Inputs):

    try:
        v = get_values(inputs)

        angle = ygol_mesta(v["gradys_d"], v["gradys_sh"], v["tochka"])
        az = azimut(v["gradys_d"], v["gradys_sh"], v["tochka"])
        dist = dalnost(v["gradys_d"], v["gradys_sh"], v["tochka"])

        text = f"""
Результаты вычислений

Широта: {v["gradys_sh"]}
Долгота: {v["gradys_d"]}
Точка стояния: {v["tochka"]}

Угол места: {angle:.2f}°
Азимут: {az:.2f}°
Дальность: {dist:.2f} км
"""

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop, "output.docx")

        doc = Document()
        doc.add_paragraph(text)
        doc.save(file_path)

        show_notification(page, "Word файл сохранён на рабочий стол ✅")

    except ValueError:
        show_notification(page, "Ошибка данных", True)


# =====================================================
# EXCEL
# =====================================================

def save_result_excel(page: ft.Page, inputs: Inputs):

    try:
        v = get_values(inputs)

        dist = dalnost(v["gradys_d"], v["gradys_sh"], v["tochka"])
        loss = round(poteri(v["chastota"], v["diametr"], v["k"], v["tochka"]), 3)

        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Параметр"
        ws["B1"] = "Значение"

        ws["A2"] = "Дальность"
        ws["B2"] = dist

        ws["A3"] = "Потери"
        ws["B3"] = loss

        for cell in ["A1", "B1", "A2", "B2", "A3", "B3"]:
            ws[cell].alignment = Alignment(horizontal="center")

        wb.save("Результаты_Расчётов.xlsx")

        show_notification(page, "Excel файл сохранён ✅")

    except ValueError:
        show_notification(page, "Ошибка данных", True)


# =====================================================
# MENU BUTTON
# =====================================================

def create_save_menu(page: ft.Page, inputs: Inputs):

    return ft.PopupMenuButton(
        icon=ft.Icons.SAVE,
        tooltip="Сохранить",
        items=[
            ft.PopupMenuItem(
                text="TXT",
                icon=ft.Icons.TEXT_SNIPPET,
                on_click=lambda e: save_result_txt(page, inputs),
            ),
            ft.PopupMenuItem(
                text="Word",
                icon=ft.Icons.DESCRIPTION,
                on_click=lambda e: save_result_word(page, inputs),
            ),
            ft.PopupMenuItem(
                text="Excel",
                icon=ft.Icons.TABLE_CHART,
                on_click=lambda e: save_result_excel(page, inputs),
            ),
        ],
        style=ft.ButtonStyle(
            shape=ft.RoundedRectangleBorder(radius=12),
            bgcolor=ft.Colors.WHITE,
        ),
    )